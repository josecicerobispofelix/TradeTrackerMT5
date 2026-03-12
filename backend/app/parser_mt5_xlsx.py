import datetime as dt
import re
import unicodedata
from io import BytesIO
from typing import Dict, List, Optional, Tuple

from openpyxl import load_workbook
from openpyxl.utils.datetime import from_excel

SYMBOL_KEYS = {"symbol", "ativo", "instrumento"}
TYPE_KEYS = {"type", "tipo"}
VOLUME_KEYS = {"volume", "vol", "lote", "lotes"}
PROFIT_KEYS = {"profit", "lucro"}
COMMISSION_KEYS = {"commission", "comissao"}
SWAP_KEYS = {"swap"}

TIME_KEYS = {"time", "horario"}
PRICE_KEYS = {"price", "preco"}

OPEN_TIME_KEYS = {"opentime", "timeopen", "time(open)", "horarioabertura"}
CLOSE_TIME_KEYS = {"closetime", "timeclose", "time(close)", "horariofechamento"}
OPEN_PRICE_KEYS = {"openprice", "priceopen", "price(open)", "precoabertura"}
CLOSE_PRICE_KEYS = {"closeprice", "priceclose", "price(close)", "precofechamento"}

DEAL_ID_KEYS = {"ticket", "deal", "position", "order", "id"}


def normalize_header(value) -> str:
    if value is None:
        return ""
    if not isinstance(value, str):
        value = str(value)
    text = value.strip().lower()
    text = unicodedata.normalize("NFKD", text)
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    text = re.sub(r"[^a-z0-9]+", "", text)
    return text


def parse_float(value, default: float = 0.0) -> float:
    if value is None:
        return default
    if isinstance(value, (int, float)):
        return float(value)
    if isinstance(value, str):
        text = value.strip().replace(",", ".")
        if text == "":
            return default
        try:
            return float(text)
        except ValueError:
            return default
    return default


def normalize_deal_id(value) -> Optional[str]:
    if value in (None, ""):
        return None
    if isinstance(value, bool):
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    return str(value).strip()


def parse_datetime(value) -> Optional[dt.datetime]:
    if value is None:
        return None
    if isinstance(value, dt.datetime):
        result = value
    elif isinstance(value, dt.date):
        result = dt.datetime.combine(value, dt.time.min)
    elif isinstance(value, (int, float)):
        try:
            result = from_excel(value)
        except Exception:
            return None
    elif isinstance(value, str):
        text = value.strip()
        for fmt in (
            "%Y.%m.%d %H:%M:%S",
            "%Y-%m-%d %H:%M:%S",
            "%d.%m.%Y %H:%M:%S",
            "%Y.%m.%d %H:%M",
            "%Y-%m-%d %H:%M",
        ):
            try:
                result = dt.datetime.strptime(text, fmt)
                break
            except ValueError:
                result = None
        if result is None:
            try:
                result = dt.datetime.fromisoformat(text)
            except ValueError:
                return None
    else:
        return None

    # MT5 exports are timezone-naive (use broker/client local time). We keep them
    # naive here and let the caller decide how to localize (e.g., using
    # APP_TIMEZONE in the upload route) to avoid accidental double conversions.
    return result


def _extract_inline_value(text: str) -> Optional[str]:
    if ":" in text:
        after = text.split(":", 1)[1].strip()
        if after:
            return after
    return None


def _next_cell_value(row: Tuple, start_idx: int) -> Optional[str]:
    for j in range(start_idx + 1, len(row)):
        value = row[j]
        if value not in (None, ""):
            return value
    return None


def parse_metadata(ws) -> Dict:
    meta: Dict[str, Optional[str]] = {}
    for row in ws.iter_rows(min_row=1, max_row=30, max_col=10, values_only=True):
        for idx, cell in enumerate(row):
            if not isinstance(cell, str):
                continue
            text = cell.strip()
            lower_norm = normalize_header(text)

            if ("account" in lower_norm or "conta" in lower_norm) and "account" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    next_value = _next_cell_value(row, idx)
                    if next_value is not None:
                        value = str(next_value).strip()
                meta["account"] = value

                if value:
                    account_digits = re.findall(r"\d+", value)
                    if account_digits:
                        meta["account"] = account_digits[0]
                    if "(" in value and ")" in value:
                        inside = value.split("(", 1)[1].split(")", 1)[0]
                        parts = [p.strip() for p in inside.split(",") if p.strip()]
                        if parts:
                            if len(parts[0]) == 3:
                                meta.setdefault("currency", parts[0])
                            if len(parts) > 1:
                                meta.setdefault("server", parts[1])

            if ("server" in lower_norm or "servidor" in lower_norm) and "server" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    next_value = _next_cell_value(row, idx)
                    if next_value is not None:
                        value = str(next_value).strip()
                meta["server"] = value

            if ("currency" in lower_norm or "moeda" in lower_norm) and "currency" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    next_value = _next_cell_value(row, idx)
                    if next_value is not None:
                        value = str(next_value).strip()
                meta["currency"] = value

            if (lower_norm.startswith("from") or lower_norm.startswith("de")) and "report_start" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    value = _next_cell_value(row, idx)
                if value:
                    parsed = parse_datetime(value)
                    meta["report_start"] = parsed.date().isoformat() if parsed else None

            if (lower_norm.startswith("to") or lower_norm.startswith("ate")) and "report_end" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    value = _next_cell_value(row, idx)
                if value:
                    parsed = parse_datetime(value)
                    meta["report_end"] = parsed.date().isoformat() if parsed else None

            if lower_norm.startswith("data") and "report_end" not in meta:
                value = _extract_inline_value(text)
                if value is None:
                    value = _next_cell_value(row, idx)
                if value:
                    parsed = parse_datetime(value)
                    meta["report_end"] = parsed.date().isoformat() if parsed else None

    if "currency" not in meta or not meta["currency"]:
        meta["currency"] = "USD"
    return meta


def find_positions_row(ws) -> Optional[int]:
    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        for cell in row:
            if isinstance(cell, str):
                normalized = normalize_header(cell)
                if "positions" in normalized or "posicoes" in normalized:
                    return i
    return None


def map_columns(headers: List[str]) -> Optional[Dict[str, int]]:
    normalized = [normalize_header(h) for h in headers]

    def find_first(keys: set) -> Optional[int]:
        for i, h in enumerate(normalized):
            if h in keys:
                return i
        return None

    open_time_idx = find_first(OPEN_TIME_KEYS)
    close_time_idx = find_first(CLOSE_TIME_KEYS)
    open_price_idx = find_first(OPEN_PRICE_KEYS)
    close_price_idx = find_first(CLOSE_PRICE_KEYS)

    time_idxs = [i for i, h in enumerate(normalized) if h in TIME_KEYS]
    price_idxs = [i for i, h in enumerate(normalized) if h in PRICE_KEYS]

    if open_time_idx is None and time_idxs:
        open_time_idx = time_idxs[0]
    if close_time_idx is None and len(time_idxs) > 1:
        close_time_idx = time_idxs[1]

    if open_price_idx is None and price_idxs:
        open_price_idx = price_idxs[0]
    if close_price_idx is None and len(price_idxs) > 1:
        close_price_idx = price_idxs[1]

    deal_id_idx = find_first(DEAL_ID_KEYS)

    symbol_idx = find_first(SYMBOL_KEYS)
    type_idx = find_first(TYPE_KEYS)
    volume_idx = find_first(VOLUME_KEYS)
    profit_idx = find_first(PROFIT_KEYS)

    if open_time_idx is None or close_time_idx is None:
        return None
    if open_price_idx is None or close_price_idx is None:
        return None
    if symbol_idx is None or type_idx is None or volume_idx is None or profit_idx is None:
        return None

    col_map = {
        "symbol": symbol_idx,
        "type": type_idx,
        "volume": volume_idx,
        "open_time": open_time_idx,
        "close_time": close_time_idx,
        "open_price": open_price_idx,
        "close_price": close_price_idx,
        "profit": profit_idx,
        "commission": find_first(COMMISSION_KEYS),
        "swap": find_first(SWAP_KEYS),
        "deal_id": deal_id_idx,
    }
    return col_map


def is_row_empty(row: Tuple) -> bool:
    for cell in row:
        if cell not in (None, ""):
            return False
    return True


def is_section_break(row: Tuple) -> bool:
    for cell in row:
        if isinstance(cell, str):
            normalized = normalize_header(cell)
            if normalized in {"orders", "positions", "ordens", "posicoes"}:
                return True
    return False


def parse_mt5_xlsx(file_bytes: bytes) -> Tuple[Dict, List[Dict]]:
    wb = load_workbook(filename=BytesIO(file_bytes), data_only=True)
    ws = wb.active

    meta = parse_metadata(ws)

    start_row = find_positions_row(ws)
    header_row = None
    col_map = None

    search_start = start_row + 1 if start_row else 1

    for idx, row in enumerate(
        ws.iter_rows(min_row=search_start, max_row=search_start + 30, values_only=True),
        start=search_start,
    ):
        if not row:
            continue
        headers = [str(cell) if cell is not None else "" for cell in row]
        mapped = map_columns(headers)
        if mapped:
            header_row = idx
            col_map = mapped
            break

    if header_row is None or col_map is None:
        for idx, row in enumerate(ws.iter_rows(values_only=True), start=1):
            headers = [str(cell) if cell is not None else "" for cell in row]
            mapped = map_columns(headers)
            if mapped:
                header_row = idx
                col_map = mapped
                break

    if header_row is None or col_map is None:
        raise ValueError("Could not locate MT5 Positions header in XLSX")

    trades: List[Dict] = []

    for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
        if is_row_empty(row) or is_section_break(row):
            break

        symbol = row[col_map["symbol"]] if col_map["symbol"] is not None else None
        trade_type = row[col_map["type"]] if col_map["type"] is not None else None
        volume = row[col_map["volume"]] if col_map["volume"] is not None else None
        open_time = row[col_map["open_time"]]
        close_time = row[col_map["close_time"]]
        open_price = row[col_map["open_price"]]
        close_price = row[col_map["close_price"]]
        profit = row[col_map["profit"]] if col_map["profit"] is not None else None
        commission = (
            row[col_map["commission"]] if col_map["commission"] is not None else 0.0
        )
        swap = row[col_map["swap"]] if col_map["swap"] is not None else 0.0
        deal_id = row[col_map["deal_id"]] if col_map["deal_id"] is not None else None

        if symbol in (None, ""):
            continue

        open_dt = parse_datetime(open_time)
        close_dt = parse_datetime(close_time)
        if not open_dt or not close_dt:
            continue

        side = str(trade_type).strip().lower() if trade_type is not None else ""
        if "buy" in side:
            side = "buy"
        elif "sell" in side:
            side = "sell"

        trade = {
            "symbol": str(symbol).strip(),
            "side": side or "unknown",
            "volume": parse_float(volume),
            "open_time": open_dt,
            "close_time": close_dt,
            "open_price": parse_float(open_price),
            "close_price": parse_float(close_price),
            "profit": parse_float(profit),
            "commission": parse_float(commission),
            "swap": parse_float(swap),
            "deal_id": normalize_deal_id(deal_id),
        }
        trades.append(trade)

    return meta, trades
